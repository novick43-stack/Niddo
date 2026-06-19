import io
import csv
import os
from datetime import datetime, timezone, date
from typing import Optional
from functools import wraps

from flask import (
    Flask, render_template, redirect, url_for,
    session, request, jsonify, send_file, Response
)
from authlib.integrations.flask_client import OAuth
from dotenv import load_dotenv
from supabase import create_client, Client

# ── Export libs ────────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# ── Config ─────────────────────────────────────────────────────────────────────
load_dotenv()

# Resolve paths relative to *this* file so the app works both locally
# and when imported from api/index.py on Vercel.
_HERE = os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    template_folder=os.path.join(_HERE, 'templates'),
    static_folder=os.path.join(_HERE, 'static'),
)
app.secret_key = os.environ['SECRET_KEY']
app.config['SESSION_COOKIE_SECURE'] = False   # True en producción (HTTPS)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10 MB

# ── Auth0 ──────────────────────────────────────────────────────────────────────
AUTH0_DOMAIN = os.environ['AUTH0_DOMAIN']

oauth = OAuth(app)
auth0 = oauth.register(
    'auth0',
    client_id=os.environ['AUTH0_CLIENT_ID'],
    client_secret=os.environ['AUTH0_CLIENT_SECRET'],
    client_kwargs={'scope': 'openid profile email'},
    server_metadata_url=f'https://{AUTH0_DOMAIN}/.well-known/openid-configuration',
)

# ── Supabase ───────────────────────────────────────────────────────────────────
supabase: Client = create_client(
    os.environ['SUPABASE_URL'],
    os.environ['SUPABASE_SERVICE_KEY'],
)

# ── Helpers ────────────────────────────────────────────────────────────────────
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def upsert_user(role: str, auth0_id: str, email: str, nombre: str) -> None:
    table = 'administradores' if role == 'admin' else 'vecinos'
    supabase.table(table).upsert(
        {'auth0_id': auth0_id, 'email': email, 'nombre': nombre, 'last_login': now_iso()},
        on_conflict='auth0_id'
    ).execute()


def get_admin_id() -> Optional[str]:
    """Devuelve el UUID de la fila en `administradores` para el usuario en sesión."""
    user = session.get('user')
    if not user:
        return None
    result = supabase.table('administradores').select('id').eq('auth0_id', user['sub']).single().execute()
    return result.data['id'] if result.data else None


def excel_response(wb: openpyxl.Workbook, filename: str) -> Response:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=filename, as_attachment=True)


def pdf_response(buf: io.BytesIO, filename: str) -> Response:
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf', download_name=filename, as_attachment=True)


def make_excel(headers: list, rows: list, sheet_name: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 14)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    return wb


def make_pdf(title: str, headers: list, rows: list) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 0.4*cm)]
    data = [headers] + rows
    col_w = (landscape(A4)[0] - 2*cm) / max(len(headers), 1)
    t = Table(data, colWidths=[col_w] * len(headers), repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7C3AED')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f0ff')]),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
    ]))
    elements.append(t)
    doc.build(elements)
    return buf


# ── Auth decorator ─────────────────────────────────────────────────────────────
def require_auth(allowed_roles=None):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user = session.get('user')
            if not user:
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify({'error': 'No autenticado'}), 401
                return redirect(url_for('login'))
            if allowed_roles and user.get('role') not in allowed_roles:
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify({'error': 'Sin permiso'}), 403
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated
    return decorator


# ── Páginas públicas ───────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/login')
def login():
    user = session.get('user')
    if user:
        return redirect(url_for('dashboard', role=user['role']))
    return render_template('login.html')


# ── Auth0 flow ─────────────────────────────────────────────────────────────────
@app.route('/auth/login')
def auth_login():
    role = request.args.get('role', 'vecino')
    if role not in ('admin', 'vecino'):
        role = 'vecino'
    session['pending_role'] = role
    callback_url = url_for('auth_callback', _external=True)
    return auth0.authorize_redirect(redirect_uri=callback_url)


@app.route('/auth/callback')
def auth_callback():
    token = auth0.authorize_access_token()
    userinfo = token.get('userinfo', {})
    auth0_id = userinfo.get('sub')
    email    = userinfo.get('email', '')
    nombre   = userinfo.get('name', email)
    role     = session.pop('pending_role', 'vecino')
    upsert_user(role, auth0_id, email, nombre)
    session['user'] = {'sub': auth0_id, 'email': email, 'name': nombre, 'role': role}
    return redirect(url_for('dashboard', role=role))


@app.route('/auth/logout')
def auth_logout():
    session.clear()
    return redirect(
        f'https://{AUTH0_DOMAIN}/v2/logout'
        f'?returnTo={url_for("index", _external=True)}'
        f'&client_id={os.environ["AUTH0_CLIENT_ID"]}'
    )


# ── Dashboards ────────────────────────────────────────────────────────────────
@app.route('/dashboard/<role>')
@require_auth()
def dashboard(role):
    user = session['user']
    if user['role'] != role:
        return redirect(url_for('dashboard', role=user['role']))
    if role == 'admin':
        return render_template('admin_dashboard.html', user=user)
    elif role == 'vecino':
        return render_template('vecino_dashboard.html', user=user)
    return redirect(url_for('login'))


# ══════════════════════════════════════════════════════════════════════════════
# API — CONSORCIOS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/consorcios', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_consorcios_list():
    admin_id = get_admin_id()
    res = supabase.table('consorcios').select('*').eq('admin_id', admin_id).order('nombre').execute()
    return jsonify(res.data)


@app.route('/api/consorcios', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_consorcios_create():
    admin_id = get_admin_id()
    d = request.json
    payload = {
        'nombre': d.get('nombre', '').strip(),
        'direccion': d.get('direccion', ''),
        'cuit': d.get('cuit', ''),
        'pisos': d.get('pisos'),
        'unidades_totales': d.get('unidades_totales'),
        'encargado_nombre': d.get('encargado_nombre', ''),
        'encargado_tel': d.get('encargado_tel', ''),
        'admin_id': admin_id,
    }
    res = supabase.table('consorcios').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/consorcios/<cid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_consorcios_update(cid):
    admin_id = get_admin_id()
    d = request.json
    payload = {k: v for k, v in {
        'nombre': d.get('nombre'),
        'direccion': d.get('direccion'),
        'cuit': d.get('cuit'),
        'pisos': d.get('pisos'),
        'unidades_totales': d.get('unidades_totales'),
        'encargado_nombre': d.get('encargado_nombre'),
        'encargado_tel': d.get('encargado_tel'),
    }.items() if v is not None}
    res = supabase.table('consorcios').update(payload).eq('id', cid).eq('admin_id', admin_id).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/consorcios/<cid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_consorcios_delete(cid):
    admin_id = get_admin_id()
    supabase.table('consorcios').delete().eq('id', cid).eq('admin_id', admin_id).execute()
    return jsonify({'ok': True})


@app.route('/api/consorcios/<cid>/unidades', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_ufs_list(cid):
    res = supabase.table('unidades_funcionales').select('*').eq('consorcio_id', cid).order('numero').execute()
    return jsonify(res.data)


@app.route('/api/consorcios/<cid>/unidades', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_ufs_create(cid):
    d = request.json
    payload = {
        'consorcio_id': cid,
        'numero': d.get('numero', '').strip(),
        'piso': d.get('piso', ''),
        'tipo': d.get('tipo', 'departamento'),
        'superficie_m2': d.get('superficie_m2'),
        'vecino_nombre': d.get('vecino_nombre', ''),
        'vecino_email': d.get('vecino_email', ''),
    }
    res = supabase.table('unidades_funcionales').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/consorcios/<cid>/unidades/<uid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_ufs_update(cid, uid):
    d = request.json
    payload = {k: v for k, v in {
        'numero': d.get('numero'),
        'piso': d.get('piso'),
        'tipo': d.get('tipo'),
        'superficie_m2': d.get('superficie_m2'),
        'vecino_nombre': d.get('vecino_nombre'),
        'vecino_email': d.get('vecino_email'),
    }.items() if v is not None}
    res = supabase.table('unidades_funcionales').update(payload).eq('id', uid).eq('consorcio_id', cid).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/consorcios/<cid>/unidades/<uid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_ufs_delete(cid, uid):
    supabase.table('unidades_funcionales').delete().eq('id', uid).eq('consorcio_id', cid).execute()
    return jsonify({'ok': True})


@app.route('/api/consorcios/<cid>/unidades/bulk', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_ufs_bulk(cid):
    """Carga masiva desde CSV. Columnas: numero,piso,tipo,superficie_m2,vecino_nombre,vecino_email"""
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No se envió archivo'}), 400
    stream = io.StringIO(file.stream.read().decode('utf-8'))
    reader = csv.DictReader(stream)
    rows = []
    for row in reader:
        rows.append({
            'consorcio_id': cid,
            'numero': row.get('numero', '').strip(),
            'piso': row.get('piso', ''),
            'tipo': row.get('tipo', 'departamento'),
            'superficie_m2': row.get('superficie_m2') or None,
            'vecino_nombre': row.get('vecino_nombre', ''),
            'vecino_email': row.get('vecino_email', ''),
        })
    if rows:
        supabase.table('unidades_funcionales').insert(rows).execute()
    return jsonify({'inserted': len(rows)})


@app.route('/api/consorcios/<cid>/export/excel')
@require_auth(allowed_roles=['admin'])
def export_consorcios_excel(cid):
    con = supabase.table('consorcios').select('*').eq('id', cid).single().execute().data
    ufs = supabase.table('unidades_funcionales').select('*').eq('consorcio_id', cid).order('numero').execute().data
    headers = ['UF', 'Piso', 'Tipo', 'Superficie m²', 'Vecino', 'Email']
    rows = [[u['numero'], u.get('piso',''), u.get('tipo',''), u.get('superficie_m2',''),
             u.get('vecino_nombre',''), u.get('vecino_email','')] for u in ufs]
    wb = make_excel(headers, rows, 'Unidades')
    return excel_response(wb, f"consorcio_{con.get('nombre','')}.xlsx")


@app.route('/api/consorcios/<cid>/export/pdf')
@require_auth(allowed_roles=['admin'])
def export_consorcios_pdf(cid):
    con = supabase.table('consorcios').select('*').eq('id', cid).single().execute().data
    ufs = supabase.table('unidades_funcionales').select('*').eq('consorcio_id', cid).order('numero').execute().data
    headers = ['UF', 'Piso', 'Tipo', 'Sup. m²', 'Vecino', 'Email']
    rows = [[u['numero'], u.get('piso',''), u.get('tipo',''), str(u.get('superficie_m2','')),
             u.get('vecino_nombre',''), u.get('vecino_email','')] for u in ufs]
    buf = make_pdf(f"Consorcio: {con.get('nombre','')}", headers, rows)
    return pdf_response(buf, f"consorcio_{con.get('nombre','')}.pdf")


# ══════════════════════════════════════════════════════════════════════════════
# API — PROVEEDORES
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/proveedores', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_proveedores_list():
    admin_id = get_admin_id()
    res = supabase.table('proveedores').select('*').eq('admin_id', admin_id).order('nombre').execute()
    return jsonify(res.data)


@app.route('/api/proveedores', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_proveedores_create():
    admin_id = get_admin_id()
    d = request.json
    payload = {
        'nombre': d.get('nombre', '').strip(),
        'cuit': d.get('cuit', ''),
        'rubro': d.get('rubro', ''),
        'email': d.get('email', ''),
        'telefono': d.get('telefono', ''),
        'admin_id': admin_id,
    }
    res = supabase.table('proveedores').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/proveedores/<pid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_proveedores_update(pid):
    admin_id = get_admin_id()
    d = request.json
    payload = {k: v for k, v in d.items() if k in ('nombre','cuit','rubro','email','telefono') and v is not None}
    res = supabase.table('proveedores').update(payload).eq('id', pid).eq('admin_id', admin_id).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/proveedores/<pid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_proveedores_delete(pid):
    admin_id = get_admin_id()
    supabase.table('proveedores').delete().eq('id', pid).eq('admin_id', admin_id).execute()
    return jsonify({'ok': True})


@app.route('/api/proveedores/<pid>/gastos')
@require_auth(allowed_roles=['admin'])
def api_proveedores_gastos(pid):
    res = supabase.table('gastos').select('*, consorcios(nombre)').eq('proveedor_id', pid).order('fecha_gasto', desc=True).execute()
    return jsonify(res.data)


# ══════════════════════════════════════════════════════════════════════════════
# API — GASTOS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/gastos', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_gastos_list():
    admin_id = get_admin_id()
    q = supabase.table('gastos').select('*, consorcios(nombre), proveedores(nombre)').eq('admin_id', admin_id)
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', request.args['consorcio_id'])
    if request.args.get('desde'):
        q = q.gte('fecha_gasto', request.args['desde'])
    if request.args.get('hasta'):
        q = q.lte('fecha_gasto', request.args['hasta'])
    res = q.order('fecha_gasto', desc=True).execute()
    return jsonify(res.data)


@app.route('/api/gastos', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_gastos_create():
    admin_id = get_admin_id()
    d = request.json
    payload = {
        'consorcio_id': d['consorcio_id'],
        'proveedor_id': d.get('proveedor_id') or None,
        'descripcion': d.get('descripcion', '').strip(),
        'categoria': d.get('categoria', ''),
        'monto': d.get('monto', 0),
        'fecha_gasto': d.get('fecha_gasto', str(date.today())),
        'fecha_vencimiento': d.get('fecha_vencimiento') or None,
        'pagado': d.get('pagado', False),
        'fecha_pago': d.get('fecha_pago') or None,
        'metodo_pago': d.get('metodo_pago', ''),
        'recurrente': d.get('recurrente', False),
        'frecuencia': d.get('frecuencia', ''),
        'notas': d.get('notas', ''),
        'admin_id': admin_id,
    }
    res = supabase.table('gastos').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/gastos/<gid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_gastos_update(gid):
    admin_id = get_admin_id()
    d = request.json
    allowed = ('consorcio_id','proveedor_id','descripcion','categoria','monto','fecha_gasto',
                'fecha_vencimiento','pagado','fecha_pago','metodo_pago','recurrente','frecuencia','notas')
    payload = {k: v for k, v in d.items() if k in allowed}
    res = supabase.table('gastos').update(payload).eq('id', gid).eq('admin_id', admin_id).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/gastos/<gid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_gastos_delete(gid):
    admin_id = get_admin_id()
    supabase.table('gastos').delete().eq('id', gid).eq('admin_id', admin_id).execute()
    return jsonify({'ok': True})


@app.route('/api/gastos/export')
@require_auth(allowed_roles=['admin'])
def api_gastos_export():
    admin_id = get_admin_id()
    q = supabase.table('gastos').select('*, consorcios(nombre), proveedores(nombre)').eq('admin_id', admin_id)
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', request.args['consorcio_id'])
    if request.args.get('desde'):
        q = q.gte('fecha_gasto', request.args['desde'])
    if request.args.get('hasta'):
        q = q.lte('fecha_gasto', request.args['hasta'])
    data = q.order('fecha_gasto', desc=True).execute().data

    headers = ['Fecha', 'Consorcio', 'Descripción', 'Categoría', 'Proveedor', 'Monto', 'Pagado', 'Método Pago', 'Recurrente']
    rows = [[
        g.get('fecha_gasto',''), (g.get('consorcios') or {}).get('nombre',''),
        g.get('descripcion',''), g.get('categoria',''),
        (g.get('proveedores') or {}).get('nombre',''),
        g.get('monto',0), 'Sí' if g.get('pagado') else 'No',
        g.get('metodo_pago',''), 'Sí' if g.get('recurrente') else 'No'
    ] for g in data]

    fmt = request.args.get('fmt', 'excel')
    if fmt == 'pdf':
        buf = make_pdf('Historial de Gastos', headers, [list(map(str, r)) for r in rows])
        return pdf_response(buf, 'gastos.pdf')
    wb = make_excel(headers, rows, 'Gastos')
    return excel_response(wb, 'gastos.xlsx')


# ══════════════════════════════════════════════════════════════════════════════
# API — COBROS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/cobros', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_cobros_list():
    q = supabase.table('cobros').select('*, unidades_funcionales(numero, vecino_nombre, vecino_email), consorcios(nombre)')
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', request.args['consorcio_id'])
    if request.args.get('periodo'):
        q = q.eq('periodo', request.args['periodo'])
    if request.args.get('estado'):
        q = q.eq('estado', request.args['estado'])
    res = q.order('created_at', desc=True).execute()
    return jsonify(res.data)


@app.route('/api/cobros/generar', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_cobros_generar():
    """Genera un cobro para cada UF del consorcio en el período dado."""
    d = request.json
    consorcio_id = d['consorcio_id']
    periodo = d['periodo']
    monto_base = float(d['monto_base'])
    fecha_vencimiento = d.get('fecha_vencimiento')

    # Traer todas las UFs del consorcio
    ufs = supabase.table('unidades_funcionales').select('id').eq('consorcio_id', consorcio_id).execute().data

    rows = []
    for uf in ufs:
        rows.append({
            'unidad_id': uf['id'],
            'consorcio_id': consorcio_id,
            'periodo': periodo,
            'monto_base': monto_base,
            'interes_mora': 0,
            'total': monto_base,
            'estado': 'pendiente',
            'fecha_vencimiento': fecha_vencimiento,
        })
    if rows:
        supabase.table('cobros').insert(rows).execute()
    return jsonify({'generados': len(rows)})


@app.route('/api/cobros/<rid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_cobros_update(rid):
    d = request.json
    allowed = ('estado','fecha_pago','interes_mora','total','notas','comprobante_nombre')
    payload = {k: v for k, v in d.items() if k in allowed}
    res = supabase.table('cobros').update(payload).eq('id', rid).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/cobros/mora')
@require_auth(allowed_roles=['admin'])
def api_cobros_mora():
    """Cobros vencidos o en mora para el panel de morosidad."""
    q = supabase.table('cobros').select('*, unidades_funcionales(numero, vecino_nombre, vecino_email), consorcios(nombre)')
    q = q.in_('estado', ['vencido', 'en_mora'])
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', request.args['consorcio_id'])
    res = q.order('fecha_vencimiento').execute()
    return jsonify(res.data)


@app.route('/api/cobros/export')
@require_auth(allowed_roles=['admin'])
def api_cobros_export():
    q = supabase.table('cobros').select('*, unidades_funcionales(numero, vecino_nombre), consorcios(nombre)')
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', request.args['consorcio_id'])
    if request.args.get('periodo'):
        q = q.eq('periodo', request.args['periodo'])
    data = q.order('created_at', desc=True).execute().data

    headers = ['Consorcio', 'UF', 'Vecino', 'Período', 'Monto Base', 'Interés', 'Total', 'Estado', 'Vencimiento', 'Fecha Pago']
    rows = [[
        (c.get('consorcios') or {}).get('nombre',''),
        (c.get('unidades_funcionales') or {}).get('numero',''),
        (c.get('unidades_funcionales') or {}).get('vecino_nombre',''),
        c.get('periodo',''), c.get('monto_base',0), c.get('interes_mora',0),
        c.get('total',0), c.get('estado',''),
        c.get('fecha_vencimiento',''), c.get('fecha_pago',''),
    ] for c in data]

    fmt = request.args.get('fmt', 'excel')
    if fmt == 'pdf':
        buf = make_pdf('Cobros / Expensas', headers, [list(map(str, r)) for r in rows])
        return pdf_response(buf, 'cobros.pdf')
    wb = make_excel(headers, rows, 'Cobros')
    return excel_response(wb, 'cobros.xlsx')


# ══════════════════════════════════════════════════════════════════════════════
# API — BALANCE
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/balance')
@require_auth(allowed_roles=['admin'])
def api_balance():
    admin_id = get_admin_id()
    consorcio_id = request.args.get('consorcio_id')
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')

    # Ingresos (cobros pagados)
    q_cobros = supabase.table('cobros').select('total, periodo, consorcios(nombre)')
    q_cobros = q_cobros.eq('estado', 'pagado')
    if consorcio_id:
        q_cobros = q_cobros.eq('consorcio_id', consorcio_id)
    if desde:
        q_cobros = q_cobros.gte('fecha_pago', desde)
    if hasta:
        q_cobros = q_cobros.lte('fecha_pago', hasta)
    cobros = q_cobros.execute().data

    # Egresos (gastos)
    q_gastos = supabase.table('gastos').select('monto, fecha_gasto, descripcion, categoria, consorcios(nombre)').eq('admin_id', admin_id)
    if consorcio_id:
        q_gastos = q_gastos.eq('consorcio_id', consorcio_id)
    if desde:
        q_gastos = q_gastos.gte('fecha_gasto', desde)
    if hasta:
        q_gastos = q_gastos.lte('fecha_gasto', hasta)
    gastos = q_gastos.execute().data

    total_ingresos = sum(c.get('total', 0) or 0 for c in cobros)
    total_egresos  = sum(g.get('monto', 0) or 0 for g in gastos)

    return jsonify({
        'ingresos': total_ingresos,
        'egresos': total_egresos,
        'resultado': total_ingresos - total_egresos,
        'cobros': cobros,
        'gastos': gastos,
    })


@app.route('/api/balance/export')
@require_auth(allowed_roles=['admin'])
def api_balance_export():
    admin_id = get_admin_id()
    consorcio_id = request.args.get('consorcio_id')
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')

    q_cobros = supabase.table('cobros').select('total, periodo, consorcios(nombre)').eq('estado', 'pagado')
    if consorcio_id: q_cobros = q_cobros.eq('consorcio_id', consorcio_id)
    if desde: q_cobros = q_cobros.gte('fecha_pago', desde)
    if hasta: q_cobros = q_cobros.lte('fecha_pago', hasta)
    cobros = q_cobros.execute().data

    q_gastos = supabase.table('gastos').select('monto, fecha_gasto, descripcion, categoria, consorcios(nombre)').eq('admin_id', admin_id)
    if consorcio_id: q_gastos = q_gastos.eq('consorcio_id', consorcio_id)
    if desde: q_gastos = q_gastos.gte('fecha_gasto', desde)
    if hasta: q_gastos = q_gastos.lte('fecha_gasto', hasta)
    gastos = q_gastos.execute().data

    headers = ['Tipo', 'Consorcio', 'Descripción/Período', 'Categoría', 'Monto']
    rows = []
    for c in cobros:
        rows.append(['INGRESO', (c.get('consorcios') or {}).get('nombre',''), c.get('periodo',''), 'Expensas', str(c.get('total',0))])
    for g in gastos:
        rows.append(['EGRESO', (g.get('consorcios') or {}).get('nombre',''), g.get('descripcion',''), g.get('categoria',''), str(g.get('monto',0))])

    fmt = request.args.get('fmt', 'excel')
    if fmt == 'pdf':
        buf = make_pdf('Balance Financiero', headers, rows)
        return pdf_response(buf, 'balance.pdf')
    wb = make_excel(headers, rows, 'Balance')
    return excel_response(wb, 'balance.xlsx')


# ── API: datos del dashboard ───────────────────────────────────────────────────
@app.route('/api/dashboard/kpis')
@require_auth(allowed_roles=['admin'])
def api_dashboard_kpis():
    admin_id = get_admin_id()
    consorcios_count = len(supabase.table('consorcios').select('id').eq('admin_id', admin_id).execute().data)
    gastos_pendientes = len(supabase.table('gastos').select('id').eq('admin_id', admin_id).eq('pagado', False).execute().data)
    mora_count = len(supabase.table('cobros').select('id').in_('estado', ['vencido','en_mora']).execute().data)
    return jsonify({'consorcios': consorcios_count, 'gastos_pendientes': gastos_pendientes, 'en_mora': mora_count})


# ── API: perfil ────────────────────────────────────────────────────────────────
@app.route('/api/me')
@require_auth()
def api_me():
    user = session['user']
    table = 'administradores' if user['role'] == 'admin' else 'vecinos'
    result = supabase.table(table).select('*').eq('auth0_id', user['sub']).single().execute()
    return jsonify(result.data)


# ── Run ────────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print('🏢 Niddo server starting...')
    print('📍 http://localhost:3500')
    app.run(host='127.0.0.1', port=3500, debug=True)
